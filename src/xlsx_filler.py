import pandas as pd
import json
import os
from openpyxl import load_workbook

def fill_excel_template(json_path, template_path, output_path):
    # 读取JSON数据
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 加载Excel模板
    wb = load_workbook(template_path)
    ws = wb.active
    
    # 定义映射关系（根据实际模板调整）
    mappings = {
        '名称': {
            'range': 'B5:E5',  # 横向合并单元格
            'merge': True      # 表示需要合并
        },
        '类型': {
            'range':'J15:K15',
            'merge': True
        },
        '注册资本': {
            'range':'E15:F15',
            'merge': True
        },
        '法定代表人': {
            'range': 'H15',
            'merge': False
        },
        '成立日期': {
            'range':'B15:C15',
            'merge': True
        },
        '住所': {
            'range':'B16:F16',
            'merge': True
        }
    }
    
    # 填充数据
    for key, config in mappings.items():
        if key in data:
            if config['merge']:
                # 处理合并单元格
                start_cell, end_cell = config['range'].split(':')
                ws[start_cell] = data[key]
                if start_cell != end_cell:
                    ws.merge_cells(config['range'])
            else:
                # 处理单个单元格
                ws[config['range']] = data[key]
    
    # 保存结果
    wb.save(output_path)

def process_business_license():
    # 设置文件路径
    json_path = 'output/营业执照.json'
    template_path = 'input/excel-template.xlsx'
    output_path = 'output/filled_template.xlsx'
    
    # 执行填充
    fill_excel_template(json_path, template_path, output_path)

# 添加以下代码来执行程序
if __name__ == '__main__':
    process_business_license()